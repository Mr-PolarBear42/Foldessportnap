from flask import Flask, redirect, render_template, request, jsonify
import openpyxl
from openpyxl import load_workbook
from flask import send_file

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('login.html')

@app.route('/download/<filename>')
def download_file(filename):
    print(filename)
    if filename=="Jelentkezők.xlsx":
        file_path="C:/Users/szori/OneDrive/Asztali gép/Zalán mappa/Coding/Foldessportnap/Jelentkezők.xlsx"
    else:
        file_path = f"C:/Users/szori/OneDrive/Asztali gép/Zalán mappa/Coding/Foldessportnap/docs/{filename}"
    try:
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})
    
@app.route('/<page>')
def show_page(page):
    return render_template(f'{page}.html')

@app.route('/api/data', methods=['POST'])
def receive_data():
    data = request.get_json()  # Get JSON data from request
    print("Received data:", data)
    
    data_list = list(data.values())
    
    
    data_list[5] = str(data_list[5]) 
    data_list[5].removeprefix("{")
    data_list[5].removesuffix("}")
    
    print("Converted data to list:", data_list)
    wb=load_workbook("C:/Users/szori/OneDrive/Asztali gép/Zalán mappa/Coding/Foldessportnap/Jelentkezők.xlsx")
    print(wb.sheetnames)
    ws=wb.active
    ws.append(data_list)
    wb.save("Jelentkezők.xlsx")
    
    
    return jsonify({"status": "success", "message": "Data received"})

if __name__ == '__main__':
    app.run(debug=True, port=5000)